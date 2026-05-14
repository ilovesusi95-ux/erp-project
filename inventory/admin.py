from django.contrib import admin
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.urls import reverse
from django.utils.html import format_html
from .models import Supplier, RegistrationCertificate, Product, InboundOrder, InboundItem, Specification, SpecModel, SkuDatabase

@admin.register(Supplier)
class SupplierAdmin(admin.ModelAdmin):
    list_display = ('name', 'address', 'production_license', 'contact_phone')
    search_fields = ('name', 'production_license', 'contact_phone')


class SpecModelInline(admin.TabularInline):
    model = SpecModel
    extra = 1
    fields = ('name', 'drainage_ball', 'drainage_bag', 'puncture_needle')


class SpecificationInline(admin.TabularInline):
    model = Specification
    extra = 1
    fields = ('name',)


@admin.register(RegistrationCertificate)
class RegistrationCertificateAdmin(admin.ModelAdmin):
    list_display = ('name', 'number', 'shelf_life_months')
    search_fields = ('name', 'number')
    inlines = [SpecificationInline]


@admin.register(Specification)
class SpecificationAdmin(admin.ModelAdmin):
    list_display = ('name', 'registration')
    search_fields = ('name',)
    list_filter = ('registration',)
    inlines = [SpecModelInline]


@admin.register(SpecModel)
class SpecModelAdmin(admin.ModelAdmin):
    list_display = ('name', 'specification', 'drainage_ball', 'drainage_bag', 'puncture_needle')
    search_fields = ('name',)
    list_filter = ('specification__registration',)

    def has_module_permission(self, request):
        return False


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ('name', 'sku', 'specification', 'model', 'registration', 'price')
    search_fields = ('name', 'sku')
    list_filter = ('registration',)


class InboundItemInline(admin.TabularInline):
    model = InboundItem
    extra = 3
    fields = ('registration', 'product', 'batch_number', 'production_date', 'expiry_date', 'quantity')
    readonly_fields = ('production_date', 'expiry_date')


@admin.register(InboundOrder)
class InboundOrderAdmin(admin.ModelAdmin):
    list_display = ('order_number', 'inbound_date', 'supplier', 'item_count', 'total_quantity')
    inlines = [InboundItemInline]
    search_fields = ('order_number',)
    actions = ['download_excel']

    def item_count(self, obj):
        return obj.items.count()
    item_count.short_description = '商品种类'

    def total_quantity(self, obj):
        return sum(item.quantity for item in obj.items.all())
    total_quantity.short_description = '总数量'

    def download_excel(self, request, queryset):
        if len(queryset) != 1:
            self.message_user(request, "请只选择一条入库单", level='ERROR')
            return
        order = queryset.first()

        wb = Workbook()
        ws = wb.active
        ws.title = "入库单"

        # 标题
        ws.merge_cells('A1:O1')
        ws['A1'] = "张三科技有限公司入库单"
        ws['A1'].font = Font(size=18, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # 供应商信息
        ws.merge_cells('A3:B3')
        ws['A3'] = "供 应 商："
        ws['C3'] = str(order.supplier) if order.supplier else ""
        ws.merge_cells('A4:B4')
        ws['A4'] = "入库日期："
        ws['C4'] = order.inbound_date.strftime('%Y-%m-%d')
        ws.merge_cells('A5:B5')
        ws['A5'] = "入库单号："
        ws['C5'] = order.order_number

        # 表头
        headers = ["序号", "商品名称", "注册证号", "规格型号", "生产批号", "生产日期", "到期日期", "数量", "单位", "单价", "金额", "生产企业", "生产许可证号", "质量状况", "备注"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # 明细数据
        for row_idx, item in enumerate(order.items.all(), 8):
            ws.cell(row=row_idx, column=1, value=row_idx-7).alignment = Alignment(horizontal='center')
            ws.cell(row=row_idx, column=2, value=item.product.name if item.product else "")
            ws.cell(row=row_idx, column=3, value=item.registration.number if item.registration else "")
            ws.cell(row=row_idx, column=4, value=item.product.specification if item.product else "")
            ws.cell(row=row_idx, column=5, value=item.batch_number)
            ws.cell(row=row_idx, column=6, value=item.production_date.strftime('%Y-%m-%d') if item.production_date else "")
            ws.cell(row=row_idx, column=7, value=item.expiry_date.strftime('%Y-%m-%d') if item.expiry_date else "")
            ws.cell(row=row_idx, column=8, value=item.quantity)
            ws.cell(row=row_idx, column=9, value="套")
            ws.cell(row=row_idx, column=10, value="")
            ws.cell(row=row_idx, column=11, value="")
            ws.cell(row=row_idx, column=12, value=item.registration.name if item.registration else "")
            ws.cell(row=row_idx, column=13, value="")
            ws.cell(row=row_idx, column=14, value="合格")
            ws.cell(row=row_idx, column=15, value="")

        # 合计
        total_qty = sum(item.quantity for item in order.items.all())
        ws.merge_cells('A12:G12')
        ws['A12'] = f"合计（大写）：伍千零伍百元整"
        ws['A12'].font = Font(bold=True)
        ws['H12'] = total_qty

        # 页脚
        ws.merge_cells('A14:O14')
        ws['A14'] = "公司地址：北京市大兴区经济开发区科苑路18号5号楼B座5层513室          联系电话：010-80227188"
        ws['A14'].font = Font(size=9)

        ws.merge_cells('A16:O16')
        ws['A16'] = "货物如有差错或质量问题，请于两天内通知本公司。制单人：________  复核人：________  入库人：________"
        ws['A16'].font = Font(size=9)

        response = HttpResponse(content_type='application/vnd.openxmlformats.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=入库单_{order.order_number}.xlsx'
        wb.save(response)
        return response

    download_excel.short_description = "下载入库单Excel"


# ====================== 新增：SKU数据库管理 ======================
@admin.register(SkuDatabase)
class SkuDatabaseAdmin(admin.ModelAdmin):
    list_display = ('specification', 'tube_type', 'drainage_ball', 'drainage_bag', 'puncture_needle', 'full_name', 'sku_code')
    list_filter = ('specification', 'tube_type', 'drainage_ball', 'drainage_bag', 'puncture_needle')
    search_fields = ('full_name', 'sku_code')
    ordering = ('specification', 'tube_type')

    actions = ['generate_all_skus']

    def generate_all_skus(self, request, queryset):
        specs = ['Fr10', 'Fr12', 'Fr14', 'Fr16', 'Fr18', 'Fr20', 'Fr22', 'Fr24']
        tubes = ['圆管', '十字管', '双腔管']
        balls = ['100ml引流球', '200ml引流球']
        bags = ['无引流袋', '700ml引流袋']
        needles = ['无穿刺针', '配穿刺针', '配可折弫穿刺针']

        created_count = 0
        for spec in specs:
            for tube in tubes:
                for ball in balls:
                    for bag in bags:
                        for needle in needles:
                            full_name = (
                                f"一次性使用术后引流管套件III型{spec}{tube}{ball}{bag}{needle}"
                            )
                            if not SkuDatabase.objects.filter(full_name=full_name).exists():
                                SkuDatabase.objects.create(
                                    specification=spec,
                                    tube_type=tube,
                                    drainage_ball=ball,
                                    drainage_bag=bag,
                                    puncture_needle=needle,
                                )
                                created_count += 1

        self.message_user(request, f"成功生成 {created_count} 条新SKU！总共288条组合已完成。")

    generate_all_skus.short_description = "一键生成所有 288 条 SKU 组合"
